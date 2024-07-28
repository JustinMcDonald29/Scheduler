import copy
import os
from datetime import datetime, timedelta

import networkx as nx
import pandas as pd
import openpyxl

import random
from itertools import product, combinations
from collections import Counter

from openpyxl.styles import PatternFill

FinalBestMat = None
FinalBestTeam = None
FinalBestScore = 1e50
coloring = {}
cteams = 20
slots = 10
days = 10
nights = 5
occ = []

def is_valid_coloring(G, node, color, coloring):
    for neighbor in G.neighbors(node):
        if neighbor in coloring and coloring[neighbor] == color:
            return False
    return True


def backtrack_coloring(G, coloring, nodes):
    if len(nodes) == 0:
        return True

    # Select node with fewest neighbors already colored
    node = min(nodes, key=lambda n: len([neighbor for neighbor in G.neighbors(n) if neighbor in coloring]))

    nodes.remove(node)

    for color in range(1, max(coloring.values(), default=0) + 2):
        if is_valid_coloring(G, node, color, coloring):
            coloring[node] = color
            if backtrack_coloring(G, coloring, nodes):
                return True
            del coloring[node]

    nodes.append(node)
    return False


class Team:
    def __init__(self, args):
        self.index = args[0]
        self.team_name = args[1]
        self.division = args[2]
        self.cross = args[3]
        self.availability = []
        self.nights = [0] * nights
        for arg in args[4:]:
            self.availability.append(1-arg)

        def __lt__(other):
            return self.index < other.index
class Game:
    def __init__(self,index, team1, team2):
        self.index = index
        self.team1 = team1
        self.team2 = team2
def randomize(Teams):
    random.shuffle(Teams)
    for i in range(len(Teams)):
        Teams[i].index = i
def conflict(t1, t2):
    con = 0
    for i in range(len(t1.availability)):
        for j in range(len(t2.availability)):
            con += 1 - max(t1.availability[i] , t2.availability[j])
    return 1000000 if con == len(t1.availability) else con
def combinatorial(matrix, Teams, matptr, missing):
            lindexes = [_ for _ in range(0, days)]
            for subset_size in range(2, days + 1):
                num_combinations = combinations(lindexes, subset_size)
                G_max_matching = nx.Graph()
                G_min_color = nx.Graph()
                for subset in num_combinations:
                    s = 0
                    for day in subset:
                        s += matptr[day]
                    s += len(missing)
                    if(s > (subset_size - 1) * slots):
                        continue
                    for day in subset:
                        for cell in range(0, matptr[day]):
                            G_max_matching.add_edge(matrix[day][cell][0], matrix[day][cell][1])
                            G_min_color.add_node( (matrix[day][cell][0], matrix[day][cell][1]) )
                    for job in missing:
                        G_max_matching.add_edge(job[0], job[1])
                        G_min_color.add_node((job[0], job[1]))
                    for n1 in G_min_color.nodes:
                        for n2 in G_min_color.nodes:
                            if(n1 != n2 and
                                    len({n1[0], n1[1], n2[0], n2[1]}) == 3
                            ):
                                G_min_color.add_edge(n1, n2)
                    rem = 0
                    while G_max_matching.number_of_edges() > 0:
                        b = nx.max_weight_matching(G_max_matching, maxcardinality=True)
                        G_max_matching.remove_edges_from(b)
                        rem += 1
                    print(subset_size)

                    if rem == subset_size:
                        print("rem got it")
                    elif max(nx.coloring.greedy_color(G_min_color, strategy='saturation_largest_first').values()) <= subset_size - 1:
                        print("color got it")
                    else:
                        print("failed")

def set_empty_cell(matrix, matptr, Teams):
    global occ
    occ = [0] * (cteams)
    for i in range(0, cteams):
        occ[i] = [0] * (cteams)
    for i in range(0, days):
        for j in range(0, matptr[i]):
            occ[matrix[i][j][0]][matrix[i][j][1]] += 1
            occ[matrix[i][j][1]][matrix[i][j][0]] += 1
    for day in range(0, days):
        t = set(range(0, cteams))
        for j in range(0, matptr[day]):
            t.remove(matrix[day][j][0])
            t.remove(matrix[day][j][1])
        G = nx.Graph()
        for team1 in t:
            for team2 in t:
                if(team1 != team2 and Teams[team1].division == Teams[team2].division and Teams[team1].index < Teams[team2].index):
                    G.add_weighted_edges_from([(team1,
                                              team2,
                                              (1000000 - 10*occ[team1][team2]*occ[team1][team2] - conflict(Teams[team1], Teams[team2])))])
        match = nx.max_weight_matching(G, maxcardinality=True)
        for key, value in match:
            matrix[day][matptr[day]] = key, value
            matptr[day] += 1
            occ[key][value] += 1
            occ[value][key] += 1

def can_pair(Teams, t1, t2):
    return Teams[t1].division == Teams[t2].division or (Teams[t1].division == Teams[t2].cross and Teams[t1].cross == Teams[t2].division)
def rebalance_within(matrix, matptr, Teams):
    for day in range(0, days):
        for cell1 in range(0, matptr[day]):
            for cell2 in range(cell1+1, matptr[day]):
                #check if they are compatible
                #check if both occur more than once in the whole thing
                a = matrix[day][cell1]
                b = matrix[day][cell2]
                if(occ[a[0]][a[1]] < occ[b[0]][b[1]]):
                    a, b = b, a
                if occ[b[0]][b[1]] < 2 or occ[b[0]][b[1]] == occ[a[0]][a[1]]:
                    continue
                bscore = occ[a[0]][a[1]] ** 2 + occ[b[0]][b[1]] ** 2

                bteams = 'b'
                #ac
                if occ[a[0]][b[0]] ** 2 + occ[b[1]][a[1]] ** 2 < bscore and (can_pair(Teams, a[0], b[0]) and can_pair(Teams, a[1], b[1])):
                    bteams = 'c'
                    bscore = occ[a[0]][b[0]] ** 2 + occ[b[1]][a[1]] ** 2
                #ad
                if occ[a[0]][b[1]] ** 2 + occ[b[0]][a[1]] ** 2 < bscore and (can_pair(Teams, a[0], b[1]) and can_pair(Teams, a[1], b[0])):
                    bteams = 'd'
                    bscore = occ[a[0]][b[1]] ** 2 + occ[b[0]][a[1]] ** 2
                if bteams == 'c':
                    g1 = (a[0], b[0])
                    g2 = (a[1], b[1])
                    occ[a[0]][a[1]] -= 1
                    occ[a[1]][a[0]] -= 1
                    occ[b[0]][b[1]] -= 1
                    occ[b[1]][b[0]] -= 1

                    occ[a[0]][b[0]] += 1
                    occ[a[1]][b[1]] += 1
                    occ[b[0]][a[0]] += 1
                    occ[b[1]][a[1]] += 1

                    matrix[day][cell1] = g1
                    matrix[day][cell2] = g2

                if bteams == 'd':
                    g1 = (a[0], b[1])
                    g2 = (a[1], b[0])
                    occ[a[0]][a[1]] -= 1
                    occ[a[1]][a[0]] -= 1
                    occ[b[0]][b[1]] -= 1
                    occ[b[1]][b[0]] -= 1

                    occ[a[0]][b[1]] += 1
                    occ[a[1]][b[0]] += 1
                    occ[b[0]][a[1]] += 1
                    occ[b[1]][a[0]] += 1

                    matrix[day][cell1] = g1
                    matrix[day][cell2] = g2

def schedule(matrix, matptr, Teams):
    #weight = (100000000 - sum(timescelltaken - mintimecelltaken) ) * avilability + 1
    for d in range(0, days):
        G = nx.Graph()
        vector = copy.deepcopy(matrix[d])
        for cell in range(0, matptr[d]):
            availability = []
            t1 = Teams[matrix[d][cell][0]]
            t2 = Teams[matrix[d][cell][1]]
            for i in range(len(Teams[matrix[d][cell][0]].availability)):
                availability.append(min(Teams[matrix[d][cell][0]].availability[i], Teams[matrix[d][cell][1]].availability[i]))
            for slot in range(0,  matptr[d]):
                G.add_weighted_edges_from([(1000*(cell + 1), slot, 100003 if availability[slot]==0 else 100000000000000 - (t1.nights[slot % nights] + t2.nights[slot % nights])**4)])
        blossom = nx.max_weight_matching(G, maxcardinality=True)
        for key, value in blossom:
            job = max(key, value) // 1000 - 1
            s = min(key, value)
            matrix[d][s] = copy.deepcopy(vector[job])
            t1 = Teams[matrix[d][s][0]]
            t2 = Teams[matrix[d][s][1]]
            t1.nights[s % nights] += 1
            t2.nights[s % nights] += 1

def score_table(matrix, matptr, Teams):
    score = 0 #less is better
    o = [[0] * cteams for __ in range(cteams)]

    for d in range(0, days):
        for cell in range(0, matptr[d]):
            o[matrix[d][cell][0]][matrix[d][cell][1]] += 1
            o[matrix[d][cell][1]][matrix[d][cell][0]] += 1
    for t1 in Teams:
        for t2 in Teams:
            if(t2.index <= t1.index):
                continue
            if(can_pair(Teams, t1.index, t2.index) and o[t1.index][t2.index] == 0):
                score += 1000000000000
    for d in range(0, days):
        for cell in range(0, matptr[d]):
            if(Teams[matrix[d][cell][0]].availability[cell] == 0 or Teams[matrix[d][cell][1]].availability[cell] == 0):
                score += 100000
    for t in Teams:
         score += (max(t.nights) - min(t.nights))**4
    return score
def heuristic_cross_first(Teams):
    for tt in range(1, 25):
        randomize(Teams)
        bestmat = 0
        mostadded = 0
        bestmatptr = 0
        bestteam = 0
        bestmissing = 0
        for i in range(1, 50):
            missing = []
            randomize(Teams)
            madded = 0
            G2 = nx.Graph()
            for t in range(len(Teams)):
                G2.add_node(t)
            for node1 in G2.nodes():
                for node2 in G2.nodes():
                    if node1 != node2:
                        if (
                                (Teams[node1].division == Teams[node2].cross and
                                 Teams[node1].cross == Teams[node2].division)
                        ):
                            G2.add_edge(node1, node2)
            p = 0
            matrix = [[None] * slots for _ in range(days)]
            matptr = [0 * slots for _ in range(days)]
            while(G2.number_of_edges() > 0):
                cross_match = nx.max_weight_matching(G2)
                for edge in cross_match:
                    matrix[p][matptr[p]] = edge
                    matptr[p] += 1
                    madded += 1
                p += 1
                G2.remove_edges_from(cross_match)

            for node1 in G2.nodes():
                for node2 in G2.nodes():
                    if node1 != node2:
                        if (
                                (Teams[node1].division == Teams[node2].division) and
                                (Teams[node1].index > Teams[node2].index)):
                            G2.add_edge(node1, node2)
            for i in range(0, slots):
                bm = sorted(nx.max_weight_matching(G2))
                G2.remove_edges_from(bm)
            tt1 = [i for i in range(0, cteams)]
            tt2 = [i for i in range(0, cteams)]
            random.shuffle(tt1)
            random.shuffle(tt2)
            for ttt1 in tt1:
                for ttt2 in tt2:
                    team1 = Teams[ttt1]
                    team2 = Teams[ttt2]
                    if(team1.division == team2.division and team1.index < team2.index):
                        bindex = -1
                        busy = -1
                        for i in range(0, days):
                            bad = matptr[i] >= slots
                            for j in range(0, matptr[i]):
                                if matrix[i][j][0] == team1.index or matrix[i][j][0] == team2.index or matrix[i][j][1] == team1.index or matrix[i][j][1] == team2.index:
                                    bad = 1
                            if not bad and matptr[i] > busy:
                                busy = matptr[i]
                                bindex = i
                        if(not bindex == -1):
                            matrix[bindex][matptr[bindex]] = (team1.index, team2.index)
                            matptr[bindex] += 1
                            madded += 1
                        else:
                            missing.append((ttt1, ttt2))

            if(madded > mostadded):
                mostadded = madded
                bestmat = copy.deepcopy(matrix)
                bestmatptr = copy.deepcopy(matptr)
                bestteam = copy.deepcopy(Teams)
                bestmissing = copy.deepcopy(missing)
        for _ in range(1, 5):
            randomize(Teams)
            madded = 0
            missing = []
            G2 = nx.Graph()
            for t in range(len(Teams)):
                G2.add_node(t)
            for node1 in G2.nodes():
                for node2 in G2.nodes():
                    if node1 != node2:
                        if (
                                (Teams[node1].division == Teams[node2].cross and
                                 Teams[node1].cross == Teams[node2].division)
                        ):
                            G2.add_edge(node1, node2)
            p = 0
            matrix = [[None] * slots for _ in range(days)]
            matptr = [0 * slots for _ in range(days)]
            while(G2.number_of_edges() > 0):
                cross_match = nx.max_weight_matching(G2)
                for edge in cross_match:
                    matrix[p][matptr[p]] = edge
                    matptr[p] += 1
                    madded += 1
                p += 1
                G2.remove_edges_from(cross_match)

            for node1 in G2.nodes():
                for node2 in G2.nodes():
                    if node1 != node2:
                        if (
                                (Teams[node1].division == Teams[node2].division) and
                                (Teams[node1].index > Teams[node2].index)):
                            G2.add_edge(node1, node2)
            for i in range(0, slots):
                bm = sorted(nx.max_weight_matching(G2))
                for team1, team2 in bm:
                    added = False
                    for day in range(days - 1, -1, -1):
                        bad = False
                        for cell in range(matptr[day]):
                            bad = (bad or
                                   matrix[day][cell][0] == team1 or
                                   matrix[day][cell][0] == team2 or
                                   matrix[day][cell][1] == team1 or
                                   matrix[day][cell][1] == team2)
                        if not bad:
                            added = True
                            matrix[day][matptr[day]] = (team1, team2)
                            matptr[day] += 1
                            madded += 1
                        if added:
                            break

                    if not added:
                        missing.append((team1, team2))
                G2.remove_edges_from(bm)
            if(madded > mostadded):
                mostadded = madded
                bestmat = copy.deepcopy(matrix)
                bestmatptr = copy.deepcopy(matptr)
                bestteam = copy.deepcopy(Teams)
                bestmissing = copy.deepcopy(missing)
        for _ in range(1, 10):
                    randomize(Teams)
                    madded = 0
                    missing = []
                    G3 = nx.Graph()
                    for t in range(len(Teams)):
                        G3.add_node(t)
                    for node1 in G3.nodes():
                        for node2 in G3.nodes():
                            if node1 != node2:
                                if (
                                        (Teams[node1].division == Teams[node2].cross and
                                         Teams[node1].cross == Teams[node2].division)
                                ):
                                    G3.add_edge(node1, node2)
                    p = 0
                    matrix = [[None] * slots for _ in range(days)]
                    matptr = [0 * slots for _ in range(days)]
                    while (G3.number_of_edges() > 0):
                        cross_match = nx.max_weight_matching(G3)
                        for edge in cross_match:
                            matrix[p][matptr[p]] = edge
                            matptr[p] += 1
                            madded += 1
                        p += 1
                        G3.remove_edges_from(cross_match)

                    for node1 in G3.nodes():
                        for node2 in G3.nodes():
                            if node1 != node2:
                                if (
                                        (Teams[node1].division == Teams[node2].division) and
                                        (Teams[node1].index > Teams[node2].index)):
                                    G3.add_edge(node1, node2)
                    for i in range(0, slots):
                        bm = sorted(nx.max_weight_matching(G3))
                        for team1, team2 in bm:
                            added = False
                            for day in range(0, days):
                                bad = False
                                for cell in range(matptr[day]):
                                    bad = (bad or
                                           matrix[day][cell][0] == team1 or
                                           matrix[day][cell][0] == team2 or
                                           matrix[day][cell][1] == team1 or
                                           matrix[day][cell][1] == team2)
                                if not bad:
                                    added = True
                                    matrix[day][matptr[day]] = (team1, team2)
                                    matptr[day] += 1
                                    madded += 1
                                if added:
                                    break

                            if not added:
                                missing.append((team1, team2))
                        G3.remove_edges_from(bm)
                    if (madded > mostadded):
                        mostadded = madded
                        bestmat = copy.deepcopy(matrix)
                        bestmatptr = copy.deepcopy(matptr)
                        bestteam = copy.deepcopy(Teams)
                        bestmissing = copy.deepcopy(missing)
        #print(bestmat)


        matrix = bestmat
        matptr = bestmatptr

        set_empty_cell(bestmat, bestmatptr, bestteam)
        rebalance_within(bestmat, bestmatptr, bestteam)
        rebalance_within(bestmat, bestmatptr, bestteam)
        rebalance_within(bestmat, bestmatptr, bestteam)
        schedule(bestmat, bestmatptr, bestteam)
        rating = score_table(bestmat, bestmatptr, bestteam)
        global FinalBestTeam
        global FinalBestMat
        global FinalBestScore
        if rating < FinalBestScore:
            FinalBestTeam = copy.deepcopy(bestteam)
            FinalBestMat = copy.deepcopy(bestmat)
            FinalBestScore = rating



def use_coloring(Teams):
    Matches = []
    mindex = 0
    for i in range(0, len(Teams)):
        for j in range(i + 1, len(Teams)):
            if (Teams[i].division == Teams[j].division or (
                    Teams[i].division == Teams[j].cross and Teams[j].division == Teams[i].cross)):
                Matches.append(Game(mindex, min(i, j), max(i, j)))
                mindex += 1

    Col = None
    coloring_strategies = [
        'largest_first',
        'random_sequential',
        'smallest_last',
        'independent_set',
        'connected_sequential_bfs',
        'connected_sequential_dfs',
        'saturation_largest_first',
        'DSATUR'
    ]
    done = 0
    best_so_far = 100
    for _ in range(1, 7):
        print(done)
        G = nx.Graph()
        for game in Matches:
            G.add_node(game)

        for node1 in G.nodes():
            for node2 in G.nodes():
                if node1 != node2:
                    if (
                            node1.team1 == node2.team1 or node1.team2 == node2.team2 or node1.team1 == node2.team2 or node1.team2 == node2.team1):
                        G.add_edge(node1, node2)

        random.shuffle(Matches)
        G = nx.Graph()
        for game in Matches:
            G.add_node(game)

        for node1 in G.nodes():
            for node2 in G.nodes():
                if node1 != node2:
                    if (
                            node1.team1 == node2.team1 or node1.team2 == node2.team2 or node1.team1 == node2.team2 or node1.team2 == node2.team1):
                        G.add_edge(node1, node2)

        done += 1
        # Add nodes to the graph

        coloring = nx.coloring.greedy_color(G, strategy=coloring_strategies[done % len(coloring_strategies)])
        # Print the coloring
        # print("Minimal coloring:", coloring)
        if max(coloring.values()) < slots:
            Col = coloring
            break
        else:
            comp = list(Counter(coloring.values()).values())
            if len(comp) - 1 == slots:
                if(comp[-2] < best_so_far):
                    best_so_far = comp[-1]
                    Col = coloring

    return Col

def edges_cross_first(Teams):
    G2 = nx.Graph()
    while True:
        randomize(Teams)
        for t in range(len(Teams)):
            G2.add_node(t)
        for node1 in G2.nodes():
            for node2 in G2.nodes():
                if node1 != node2:
                    if (
                            (Teams[node1].index < Teams[node2].index and
                             Teams[node1].division == Teams[node2].division) or
                            (Teams[node1].division == Teams[node2].cross and
                             Teams[node1].cross == Teams[node2].division)
                    ):
                        G2.add_edge(node1, node2)
        # backtrack_coloring(G, coloring, list(G.nodes())
        bm = []
        for i in range(0, 10):
            bm = sorted(nx.max_weight_matching(G2))
            G2.remove_edges_from(bm)
        # print(G2)
        if G2.number_of_edges() == 0:
            break
light_colors = [''.join([hex(random.randint(120, 200))[2:].zfill(2) for _ in range(3)]) for _ in range(10)]


df = pd.read_excel("./table.xlsx", engine='openpyxl')
first_days = pd.to_datetime(df.iloc[:, -3], format='%d/%m')
cteams = len(df['Name'])
days = int(df.iloc[1, -1])
days_in_week = int(df.iloc[0, -1])
slots = (cteams + 1) // 2
nights = slots // days_in_week
first_hour = df['first hour'][0]

df.drop(df.columns[-3:], axis=1, inplace=True)
df.fillna(0, inplace=True)
df.replace("no", 1, inplace=True)
df.iloc[:, 2:] = df.iloc[:, 2:].astype(int)
df['Cross']= df['Cross'].astype(int)



Teams = []
for index, row in df.iterrows():
    addL = [index]
    for i in row.tolist():
        addL.append(i)
    t = Team(addL)

    Teams.append(t)

randomize(Teams)

res = heuristic_cross_first(Teams)

def add_hours_to_time(time_obj, hours):
    dt_with_time = datetime.combine(datetime.today(), time_obj)
    delta_hours = timedelta(hours=hours)
    result_dt = dt_with_time + delta_hours
    result_time = result_dt.time()
    return result_time


def fix_matrix():
    random.shuffle(FinalBestMat)
    matrix = FinalBestMat
    lhome = [0] * 1000
    for i in range(0, days):
        for j in range(0, slots):
            if(matrix[i][j] != None):
                if(lhome[matrix[i][j][0]] > lhome[matrix[i][j][1]]):
                    matrix[i][j] = (matrix[i][j][1], matrix[i][j][0])
                lhome[matrix[i][j][0]] += 1
def get_render_matrix():
    date_array = []
    pdate = 0
    for i in range(0, 100):
        date_array.append(first_days[i % days_in_week] + timedelta(days = 7 * (i // days_in_week)))
    matrix = FinalBestMat
    Teams = FinalBestTeam
    final_table = []
    final_color = []
    for i in range(0, days):
        for j in range(0, slots):
            text = [""] * 4
            color = [None] * 4
            text[0] = date_array[pdate].strftime("%b-%d") if j % nights == 0 else ""
            if(j % nights == 0):
                pdate += 1
            text[1] = str(add_hours_to_time(first_hour, j % nights))
            if(matrix[i][j] == None):
                text[2] = "N/A"
                color[2] = "000000"
                text[3] = "N/A"
                color[3] = "020202"
            else:
                text[2] = Teams[matrix[i][j][0]].team_name
                color[2] = light_colors[ Teams[matrix[i][j][0]].division]
                text[3] = Teams[matrix[i][j][1]].team_name
                color[3] = light_colors[Teams[matrix[i][j][1]].division]
                final_table.append((text))
                final_color.append(color)

        final_table.append([''] * 4)
        final_color.append([None]*4)
    return final_table, final_color
def render(table, colors):
    filename = 'final.xlsx'

    #delete any previous final.xlsx
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active

    start_row = 2
    start_col = 1

    for i, row in enumerate(table):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=value)
            # Check if a color is specified for the current cell
            color = colors[i][j]
            if color is not None:
                # Create a PatternFill object with the specified color
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                # Apply the fill color to the cell
                cell.fill = fill

    wb.save(filename)


fix_matrix()
table, color = get_render_matrix()
render(table, color)

