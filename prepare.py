import openpyxl
from openpyxl.utils import get_column_letter
n_teams = int(input("Enter number of teams:"))
n_dpw = int(input("Enter number days per week:"))
total_weeks = int(input("Enter number of weeks"))
file_path = 'table.xlsx'

try:
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        workbook[sheet].delete_rows(1, workbook[sheet].max_row)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet = workbook.active

# Write to a specific cell
cell = sheet['A1']
sheet['A1'] = "Name"
sheet['B1'] = "Division"
sheet['C1'] = "Cross"
n_nights = n_teams//2 + n_teams % 2
n_nights = (n_nights + n_dpw - 1) // n_dpw
workbook.save(file_path)
for day in range(0, n_dpw):
    for night in range(0, n_nights):
        cell.offset(column = 3 + day * n_nights + night).value = night+1
coloff = 3 + n_nights * n_dpw + 1
cell.offset(column = coloff, row = 0).value = 'days of first week'
cell.offset(column = coloff+1, row = 0).value = 'first hour'
cell.offset(column = coloff+2, row = 1).value = n_dpw
cell.offset(column = coloff+2, row = 2).value = total_weeks
workbook.save(file_path)